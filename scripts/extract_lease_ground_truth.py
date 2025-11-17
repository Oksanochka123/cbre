import argparse
import json
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm


def _extract_static_fields(
    ws: Worksheet, static_fields_config: dict[str, str | list[str] | dict[str, Any]]
) -> dict[str, Any]:
    result = {}
    for field, cell_ref in static_fields_config.items():
        if isinstance(cell_ref, dict):
            actual_cell_ref = None
            if "if" in cell_ref:
                condition = cell_ref["if"]
                check_cell = condition.get("check")
                contains_text = condition.get("contains")

                if check_cell and contains_text:
                    check_value = ws[check_cell].value
                    if check_value and contains_text in str(check_value):
                        actual_cell_ref = condition.get("use")

            if actual_cell_ref is None:
                actual_cell_ref = cell_ref.get("cell") or cell_ref.get("cells")

            cell_ref = actual_cell_ref

        if isinstance(cell_ref, list):
            values = [
                str_val
                for ref in cell_ref
                if (cell_value := ws[ref].value) is not None and (str_val := str(cell_value).strip())
            ]
            value = " ".join(values) if values else None
        else:
            value = ws[cell_ref].value

        result[field] = value
    return result


def _extract_tables(ws: Worksheet, tables_config: dict[str, dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    result = {}
    for table_name, table_conf in tables_config.items():
        start_row = table_conf.get("start_row", None)
        column_map = table_conf.get("columns", {})

        if not start_row or not column_map:
            continue

        table_data = []
        current_row = start_row + 1

        while True:
            primary_col_letter = next(iter(column_map.values()))
            primary_col_index = column_index_from_string(primary_col_letter)
            first_cell_val = ws.cell(row=current_row, column=primary_col_index).value

            if first_cell_val is None or str(first_cell_val).strip() == "":
                break

            row_data = {}
            for field_name, col_letter in column_map.items():
                col_idx = column_index_from_string(col_letter)
                cell_val = ws.cell(row=current_row, column=col_idx).value
                row_data[field_name] = cell_val

            table_data.append(row_data)
            current_row += 1

        result[table_name] = table_data
    return result


def extract_lease_data(excel_path: Path | str, config: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    try:
        wb = load_workbook(excel_path, data_only=True)
    except BadZipFile as e:
        print(f"Error loading Excel file {excel_path}: {e}")
        return {}

    result = {}

    for sheet_name, sheet_config in config.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        sheet_result = {
            "static_fields": _extract_static_fields(ws, sheet_config.get("static_fields", {})),
            "tables": _extract_tables(ws, sheet_config.get("tables", {})),
        }
        result[sheet_name] = sheet_result

    return result


def process_forms(input_folder: str, config_path: str, output_folder: str) -> None:
    input_path = Path(input_folder)
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)

    with open(config_path) as f:
        config = json.load(f)

    xlsm_files = list(input_path.rglob("*.xlsm"))

    for excel_file in tqdm(xlsm_files, desc="Processing files"):
        relative_path = excel_file.relative_to(input_path)
        output_file = output_path / relative_path.with_suffix(".json")
        output_file.parent.mkdir(parents=True, exist_ok=True)

        lease_data = extract_lease_data(excel_file, config)

        with open(output_file, "w") as f:
            json.dump(lease_data, f, indent=2, default=str)


def main() -> None:
    parser = argparse.ArgumentParser(description="Process lease abstract Excel files")
    parser.add_argument("input_folder", help="Input folder containing xlsm files in subfolders")
    parser.add_argument("config_path", help="Path to lease mapping config JSON file")
    parser.add_argument("output_folder", help="Output folder for parsed JSON files")
    args = parser.parse_args()

    process_forms(args.input_folder, args.config_path, args.output_folder)


if __name__ == "__main__":
    main()
